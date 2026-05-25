# -*- coding: utf-8 -*-
"""Generate Excel template from markdown structure (sheets 0-7)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "template-phan-tich-he-thong-goc-nhin-user.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2D8B5E")
WRAP = Alignment(wrap_text=True, vertical="top")
EMPTY_ROWS = 15  # blank rows for data entry


def style_header_row(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_table(ws, start_row: int, headers: list[str], data_rows: int = EMPTY_ROWS) -> int:
    ncol = len(headers)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, ncol)
    for r in range(start_row + 1, start_row + 1 + data_rows):
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).alignment = WRAP
    return start_row + 1 + data_rows


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # --- Sheet 0 ---
    ws0 = wb.create_sheet("0-Thong tin tai lieu", 0)
    ws0["A1"] = "0. Thông tin tài liệu"
    ws0["A1"].font = Font(bold=True, size=14)
    rows0 = [
        ("Mục", "Nội dung"),
        ("Người lập", ""),
        ("Ngày", ""),
        ("Môi trường tham khảo (dev / staging / prod)", ""),
    ]
    for r, (a, b) in enumerate(rows0, start=3):
        ws0.cell(row=r, column=1, value=a)
        ws0.cell(row=r, column=2, value=b)
        ws0.cell(row=r, column=1).font = Font(bold=True) if r == 3 else Font(bold=True)
    style_header_row(ws0, 3, 2)
    ws0.cell(row=4, column=1).font = Font(bold=False)
    ws0.cell(row=5, column=1).font = Font(bold=False)
    ws0.cell(row=6, column=1).font = Font(bold=False)
    set_col_widths(ws0, [42, 50])

    # --- Sheet 1 ---
    ws1 = wb.create_sheet("1-He thong lam gi", 1)
    ws1["A1"] = "1. Hệ thống dùng để làm gì?"
    ws1["A1"].font = Font(bold=True, size=14)
    add_table(
        ws1,
        3,
        ["User", "Công việc cần làm trong hệ thống", "Kết quả mong đợi"],
    )
    set_col_widths(ws1, [18, 45, 35])

    # --- Sheet 2 ---
    ws2 = wb.create_sheet("2-Ai dung he thong", 2)
    ws2["A1"] = "2. Ai dùng hệ thống?"
    ws2["A1"].font = Font(bold=True, size=14)
    add_table(
        ws2,
        3,
        ["Role", "Mô tả ngắn", "Primary user? (Có/Không. Có thì ghi vào)", "Ghi chú"],
    )
    set_col_widths(ws2, [16, 35, 38, 28])

    # --- Sheet 3 ---
    ws3 = wb.create_sheet("3-Site", 3)
    ws3["A1"] = "3. Site — Hệ có những cổng vào nào?"
    ws3["A1"].font = Font(bold=True, size=14)
    add_table(
        ws3,
        3,
        [
            "Site (VD: Admin Portal, User Portal…)",
            "Role",
            "Platforms (website, mobile app, CMS)",
            "Ghi chú",
        ],
    )
    set_col_widths(ws3, [32, 16, 38, 28])

    # --- Sheet 4 ---
    ws4 = wb.create_sheet("4-Tich hop", 4)
    ws4["A1"] = "4. Có tích hợp không"
    ws4["A1"].font = Font(bold=True, size=14)
    add_table(
        ws4,
        3,
        [
            "Tích hợp (điền tên hệ thống tích hợp)",
            "Giải thích (là hệ thống gì? Tích hợp để giải quyết điều gì? Mục đích tích hợp?)",
            "Ghi chú (hệ thống internal hay external)",
        ],
    )
    set_col_widths(ws4, [28, 55, 28])

    # --- Sheet 5 ---
    ws5 = wb.create_sheet("5-Function Screen", 5)
    ws5["A1"] = "5. Function list / Screen list — theo Role"
    ws5["A1"].font = Font(bold=True, size=14)
    ws5["A2"] = "Nên liệt kê theo từng Role. Copy sheet này hoặc thêm block Role bên dưới khi có nhiều role."
    ws5["A2"].alignment = WRAP
    ws5.merge_cells("A2:F2")

    row = 4
    ws5.cell(row=row, column=1, value="Role:").font = Font(bold=True)
    ws5.cell(row=row, column=2, value="")
    row += 2

    ws5.cell(row=row, column=1, value="5.1 Function list (Các chức năng chính)").font = Font(bold=True)
    row += 1
    row = add_table(
        ws5,
        row,
        [
            "STT",
            "Tên chức năng chính",
            "Tên chức năng nhỏ",
            "Mô tả mục đích (user làm gì)",
            "Tiến độ tìm hiểu (todo/doing/done/pending)",
        ],
    )
    row += 2

    ws5.cell(row=row, column=1, value="5.2 Screen list (Danh sách màn hình)").font = Font(bold=True)
    row += 1
    row = add_table(
        ws5,
        row,
        [
            "STT",
            "Màn hình lớn",
            "Màn hình nhỏ",
            "Mô tả các event, item",
            "Tiến độ tìm hiểu",
        ],
    )
    row += 2

    ws5.cell(row=row, column=1, value="5.3 Màn hình tiêu biểu (1-3 màn hình)").font = Font(bold=True)
    row += 1
    for i in range(1, 4):
        ws5.cell(row=row, column=1, value=f"Tên màn hình {i:02d}:")
        ws5.cell(row=row, column=1).font = Font(bold=True)
        ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1
        ws5.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=5)
        ws5.cell(
            row=row,
            column=1,
            value="Giải thích chi tiết về các item/event trên màn hình",
        )
        ws5.cell(row=row, column=1).alignment = WRAP
        row += 4

    set_col_widths(ws5, [12, 22, 22, 40, 28])

    # --- Sheet 6 ---
    ws6 = wb.create_sheet("6-Glossary", 6)
    ws6["A1"] = "6. Thuật ngữ (Glossary)"
    ws6["A1"].font = Font(bold=True, size=14)
    add_table(
        ws6,
        3,
        ["Thuật ngữ tiếng Nhật", "Tiếng Việt", "Ghi chú/ví dụ (user-facing / nghiệp vụ)"],
    )
    set_col_widths(ws6, [24, 24, 45])

    # --- Sheet 7 ---
    ws7 = wb.create_sheet("7-Diem can xac nhan", 7)
    ws7["A1"] = "7. Điểm cần xác nhận"
    ws7["A1"].font = Font(bold=True, size=14)
    ws7.cell(row=3, column=1, value="STT")
    ws7.cell(row=3, column=2, value="Nội dung cần xác nhận")
    ws7.cell(row=3, column=3, value="Ghi chú / liên hệ")
    style_header_row(ws7, 3, 3)
    for r in range(4, 4 + EMPTY_ROWS):
        ws7.cell(row=r, column=1, value=r - 3)
        for c in range(2, 4):
            ws7.cell(row=r, column=c).alignment = WRAP
    set_col_widths(ws7, [6, 55, 30])

    wb.save(OUT)
    print("OK", OUT.exists(), OUT.stat().st_size if OUT.exists() else 0)


if __name__ == "__main__":
    main()
