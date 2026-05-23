#!/usr/bin/env python3
"""
Excel → JSON converter for QA test cases (TPL-QA-01-04 format, 6 envs).

Usage:
  python excel_to_json.py <excel_file.xlsx> <output_dir> [--feature <feature_name>]

Output structure:
  output_dir/
    meta.json
    [feature_name]/          ← nếu có --feature, tạo sub-folder
      {sheet_slug}.json      ← 1 sheet = 1 JSON file

Nếu không có --feature, JSON được đặt thẳng vào output_dir.
QA tự tổ chức các JSON vào folder chức năng lớn sau.

Column layout (row 10-11 headers, data from row 12):
  A=ID  B=Category  C=Name  D=Precondition  E=Steps  F=Test Data
  G=Expected result  H=Priority
  Round 1: I(BugID) J-L(ENV1) M-O(ENV2) P-R(ENV3) S-U(ENV4) V-X(ENV5) Y,AA,Z(ENV6)
  Round 2: AB(BugID) AC-AE(ENV1) AF-AH(ENV2) AI-AK(ENV3) AL-AN(ENV4) AO-AQ(ENV5) AR,AT,AS(ENV6)
  Round 3: AU(BugID) AV-AX(ENV1) AY-BA(ENV2) BB-BD(ENV3) BE-BG(ENV4) BH-BJ(ENV5) BK,BM,BL(ENV6)
  BN=Note
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

DATA_START_ROW = 12


def strip_empty(obj):
    """Đệ quy xóa các key có giá trị rỗng ('') hoặc dict/list rỗng sau khi strip."""
    if isinstance(obj, dict):
        cleaned = {k: strip_empty(v) for k, v in obj.items()}
        return {k: v for k, v in cleaned.items() if v != "" and v != {} and v != []}
    if isinstance(obj, list):
        return [strip_empty(i) for i in obj]
    return obj


def slugify(text: str) -> str:
    text = str(text).lower().strip()
    text = re.sub(r'[^\w\s-]', '', text, flags=re.UNICODE)
    text = re.sub(r'[\s_-]+', '-', text)
    text = re.sub(r'^-+|-+$', '', text)
    return text[:60] if text else "unnamed"


def fmt_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    v = str(val).strip()
    return "" if v.startswith("=") else v


def raw(ws, row, col) -> str:
    """Get cell value, skip formulas."""
    val = ws.cell(row, col).value
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.startswith("=") else s


def read_sheet(ws) -> tuple:
    """Parse one Module sheet → (module_name, creator, test_cases list)."""
    module_name = raw(ws, 1, 2)  # B1

    # Creator is typically at column after "Creator:" label in row 1
    creator = ""
    for col in range(3, 10):
        label = ws.cell(1, col).value
        if label and "Creator" in str(label):
            creator = raw(ws, 1, col + 1)
            break

    # Round env column mapping: (result_col, tester_col, date_col)
    # Note: ENV6 has date/tester swapped in the template
    R1 = [(10,11,12),(13,14,15),(16,17,18),(19,20,21),(22,23,24),(25,27,26)]
    R2 = [(29,30,31),(32,33,34),(35,36,37),(38,39,40),(41,42,43),(44,46,45)]
    R3 = [(48,49,50),(51,52,53),(54,55,56),(57,58,59),(60,61,62),(63,65,64)]

    def parse_round(row, bug_col, envs):
        env_data = {}
        for i, (rc, tc, dc) in enumerate(envs, 1):
            env_data[f"env{i}"] = {
                "result": raw(ws, row, rc),
                "tester": raw(ws, row, tc),
                "date":   fmt_date(ws.cell(row, dc).value),
            }
        return {"bug_id": raw(ws, row, bug_col), **env_data}

    test_cases = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        priority = raw(ws, row, 8)  # H
        if not priority:
            continue

        tc = {
            "category":        raw(ws, row, 2),   # B
            "name":            raw(ws, row, 3),   # C
            "precondition":    raw(ws, row, 4),   # D
            "steps":           raw(ws, row, 5),   # E
            "test_data":       raw(ws, row, 6),   # F
            "expected_result": raw(ws, row, 7),   # G
            "priority":        priority,           # H
            "note":            raw(ws, row, 66),  # BN
            "results": {
                "round1": parse_round(row, 9,  R1),
                "round2": parse_round(row, 28, R2),
                "round3": parse_round(row, 47, R3),
            },
        }
        test_cases.append(tc)

    return module_name, creator, test_cases


def excel_to_json(excel_path: str, output_dir: str, feature: str = None):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    base = Path(output_dir)

    # Determine output folder
    out_dir = base / slugify(feature) if feature else base
    out_dir.mkdir(parents=True, exist_ok=True)

    # Write project meta.json to base (not feature sub-folder)
    meta = {"project": "", "phase": ""}
    if "Cover" in wb.sheetnames:
        cover = wb["Cover"]
        for row in cover.iter_rows(min_row=1, max_row=10):
            vals = {i: (c.value or "") for i, c in enumerate(row) if c.value}
            text = " ".join(str(v) for v in vals.values())
            if "Project Name:" in text:
                cols = list(vals.values())
                for i, v in enumerate(cols):
                    if "Project Name:" in str(v) and i + 1 < len(cols):
                        meta["project"] = str(cols[i+1]).strip()
            if "Phase:" in text:
                cols = list(vals.values())
                for i, v in enumerate(cols):
                    if v == "Phase:" and i + 1 < len(cols):
                        meta["phase"] = str(cols[i+1]).strip()

    meta_path = base / "meta.json"
    if not meta_path.exists():
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        print(f"Written: {meta_path}")

    # Process each test case sheet
    test_sheets = [s for s in wb.sheetnames if s not in ("Cover", "Report")]
    if not test_sheets:
        print("No test case sheets found.")
        return

    for sheet_name in test_sheets:
        ws = wb[sheet_name]
        module_name, creator, test_cases = read_sheet(ws)

        if not test_cases:
            print(f"  {sheet_name}: no test cases, skipping.")
            continue

        data = {
            "module_name": module_name or sheet_name,
            "feature":     feature or "",
            "creator":     creator,
            "test_cases":  test_cases,
        }

        # Tìm file JSON hiện có khớp với module_name (sync mode)
        # Nếu không tìm thấy, dùng tên slugified từ sheet name
        out_path = None
        target_module = data["module_name"]
        for existing in out_dir.glob("*.json"):
            if existing.name == "meta.json":
                continue
            try:
                with open(existing, encoding="utf-8") as ef:
                    existing_data = json.load(ef)
                if existing_data.get("module_name") == target_module:
                    out_path = existing
                    break
            except (json.JSONDecodeError, KeyError):
                pass
        if out_path is None:
            out_path = out_dir / f"{slugify(sheet_name)}.json"

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(strip_empty(data), f, ensure_ascii=False, indent=2)
        print(f"  Written: {out_path}  ({len(test_cases)} TCs)")

    print(f"\nDone. Output: {out_dir}")


if __name__ == "__main__":
    args = sys.argv[1:]
    feature = None
    if "--feature" in args:
        i = args.index("--feature")
        feature = args[i + 1]
        args = [a for j, a in enumerate(args) if j != i and j != i + 1]

    if len(args) < 2:
        print("Usage: python excel_to_json.py <excel.xlsx> <output_dir> [--feature <name>]")
        sys.exit(1)

    excel_to_json(args[0], args[1], feature)
