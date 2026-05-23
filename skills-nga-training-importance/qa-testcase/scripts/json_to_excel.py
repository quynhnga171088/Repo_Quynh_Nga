#!/usr/bin/env python3
"""
JSON → Excel converter for QA test cases (TPL-QA-01-04 format, 6 envs).

Dùng assets/testcase_template.xlsx làm nền. Mỗi JSON file = 1 sheet.

Usage:
  python json_to_excel.py <json_dir> <output_excel.xlsx> [--template <file.xlsx>]

json_dir structure:
  json_dir/
    meta.json
    {feature}/
      {module}.json    ← 1 file = 1 sheet
    {module}.json      ← hoặc flat, không có feature folder
"""

import sys
import json
import re
import shutil
from pathlib import Path
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Font chuẩn theo template TPL-QA-01-04
DATA_FONT_NAME = "Times New Roman"
DATA_FONT_SIZE = 12

SCRIPT_DIR     = Path(__file__).parent
DEFAULT_TPL    = SCRIPT_DIR.parent / "assets" / "testcase_template.xlsx"
DATA_START_ROW = 12
TEMPLATE_SHEET = "Module 1"

# Column mapping
COL_ID          = 1   # A
COL_CATEGORY    = 2   # B
COL_NAME        = 3   # C
COL_PRECOND     = 4   # D
COL_STEPS       = 5   # E
COL_TESTDATA    = 6   # F
COL_EXPECTED    = 7   # G
COL_PRIORITY    = 8   # H
COL_NOTE        = 66  # BN

# Round 1 (cols 9-27): BugID=9, then 6 envs
# ENV6 has date/tester swapped (matching template)
R1_BUG = 9
R1_ENVS = [(10,11,12),(13,14,15),(16,17,18),(19,20,21),(22,23,24),(25,27,26)]
R2_BUG = 28
R2_ENVS = [(29,30,31),(32,33,34),(35,36,37),(38,39,40),(41,42,43),(44,46,45)]
R3_BUG = 47
R3_ENVS = [(48,49,50),(51,52,53),(54,55,56),(57,58,59),(60,61,62),(63,65,64)]


def slugify(text: str) -> str:
    text = str(text).lower().strip()
    text = re.sub(r'[^\w\s-]', '', text, flags=re.UNICODE)
    text = re.sub(r'[\s_-]+', '-', text)
    return text[:60] if text else "unnamed"


def update_countif(formula: str, last_row: int) -> str:
    return re.sub(
        r'([A-Z]{1,2})12:[A-Z]{1,2}\d+',
        lambda m: f"{m.group(1)}12:{m.group(1)}{last_row}",
        formula
    )


def clear_data_rows(ws):
    """Remove all rows from DATA_START_ROW onward, unmerge data merges."""
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= DATA_START_ROW:
            ws.merged_cells.remove(m)
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)


def capture_row_style(ws) -> dict:
    """Capture border from row 11 (sub-header row has borders).
    Font is always overridden to DATA_FONT_NAME/DATA_FONT_SIZE per template.
    Fill is NOT copied — data rows have white/no background."""
    styles = {}
    for col in range(1, 70):
        cell = ws.cell(DATA_START_ROW - 1, col)  # row 11 has borders
        if cell.has_style:
            styles[col] = {
                "border":     copy(cell.border),
                "number_fmt": cell.number_format,
            }
    return styles


def apply_style(cell, s: dict):
    # Always apply template font to data cells
    cell.font = Font(name=DATA_FONT_NAME, size=DATA_FONT_SIZE)
    if s.get("border"):    cell.border = copy(s["border"])
    if s.get("number_fmt"): cell.number_format = s["number_fmt"]


def write_sheet(ws, data: dict):
    """Write test case data into a sheet (rows 1-11 already in place from template copy)."""
    module_name = data.get("module_name", "")
    creator     = data.get("creator", "")
    test_cases  = data.get("test_cases", [])

    # Capture style before clearing
    row_style = capture_row_style(ws)
    clear_data_rows(ws)

    # Update metadata
    ws.cell(1, 2).value = module_name  # B1
    # Find Creator label and set value next to it
    for col in range(3, 10):
        if ws.cell(1, col).value and "Creator" in str(ws.cell(1, col).value):
            ws.cell(1, col + 1).value = creator
            break

    # Data validations
    dv_priority = DataValidation(type="list", formula1='"High, Normal, Low"',
                                  allow_blank=True, showDropDown=False)
    dv_result   = DataValidation(type="list", formula1="$L$2:$O$2",
                                  allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_priority)
    ws.add_data_validation(dv_result)

    for i, tc in enumerate(test_cases):
        r = DATA_START_ROW + i

        # Apply template style
        for col in range(1, 70):
            apply_style(ws.cell(r, col), row_style.get(col, {}))

        # ID
        ws.cell(r, COL_ID).value = i + 1
        ws.cell(r, COL_ID).alignment = Alignment(horizontal="center", vertical="center")

        # Core fields
        for col, key in [
            (COL_CATEGORY, "category"),
            (COL_NAME,     "name"),
            (COL_PRECOND,  "precondition"),
            (COL_STEPS,    "steps"),
            (COL_TESTDATA, "test_data"),
            (COL_EXPECTED, "expected_result"),
        ]:
            ws.cell(r, col).value = tc.get(key, "")
            ws.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")

        # Priority
        prio = tc.get("priority", "Normal")
        p_cell = ws.cell(r, COL_PRIORITY, prio)
        p_cell.alignment = Alignment(horizontal="center", vertical="center")
        dv_priority.add(p_cell)

        # Note
        ws.cell(r, COL_NOTE).value = tc.get("note", "")

        # Results
        results = tc.get("results", {})
        for rnd_key, bug_col, envs in [
            ("round1", R1_BUG, R1_ENVS),
            ("round2", R2_BUG, R2_ENVS),
            ("round3", R3_BUG, R3_ENVS),
        ]:
            rnd = results.get(rnd_key, {})
            ws.cell(r, bug_col).value = rnd.get("bug_id", "")

            for env_idx, (rc, tc_col, dc) in enumerate(envs, 1):
                env = rnd.get(f"env{env_idx}", {})

                res_cell = ws.cell(r, rc, env.get("result", ""))
                res_cell.alignment = Alignment(horizontal="center", vertical="center")
                dv_result.add(res_cell)

                ws.cell(r, tc_col).value = env.get("tester", "")

                date_val = env.get("date", "")
                if date_val:
                    try:
                        ws.cell(r, dc).value = datetime.strptime(date_val, "%Y-%m-%d")
                        ws.cell(r, dc).number_format = "YYYY-MM-DD"
                    except ValueError:
                        ws.cell(r, dc).value = date_val

    # Update COUNTIF formulas in rows 3-9
    last_row = DATA_START_ROW + len(test_cases) - 1
    if last_row >= DATA_START_ROW:
        for row in range(3, 10):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str) and "COUNTIF" in cell.value:
                    cell.value = update_countif(cell.value, last_row)

    ws.freeze_panes = "A12"


def update_report_sheet(wb, sheet_names: list):
    """
    Cập nhật Report sheet để tham chiếu đúng các module sheet thực tế.

    Template có Module 1 wired đúng (refs 'Module 1'!), Module 2-9 có #REF!.
    Hàm này:
    - Với module k (1..N): copy pattern từ Module 1, thay tên sheet + row numbers
    - Với module k (N+1..9): xóa các ô có sheet refs để tránh lỗi #REF!
    """
    if "Report" not in wb.sheetnames:
        return

    ws = wb["Report"]

    # Mỗi round: (module1_base_row,) — 9 module blocks × 6 env rows mỗi block
    ROUNDS_BASE = [50, 98, 146]
    MAX_MODULES = 9
    BLOCK_SIZE = 6   # 6 env rows per module block
    # Các cột chứa sheet-dependent data (cần update/clear)
    DATA_COLS = list(range(2, 22))  # C2..C21 (1-based: cols 2-21)

    for base_row in ROUNDS_BASE:
        # Đọc template block từ Module 1 (rows base_row .. base_row+5)
        tmpl_block = {}
        for rel in range(BLOCK_SIZE):
            tmpl_block[rel] = {}
            for col in DATA_COLS:
                v = ws.cell(base_row + rel, col).value
                if v is not None:
                    tmpl_block[rel][col] = v

        for k in range(MAX_MODULES):
            module_base = base_row + k * BLOCK_SIZE

            if k < len(sheet_names):
                # Generate công thức cho module k từ Module 1 pattern
                sheet_name = sheet_names[k]
                for rel in range(BLOCK_SIZE):
                    src_row = base_row + rel
                    dst_row = module_base + rel
                    for col in DATA_COLS:
                        src_val = tmpl_block[rel].get(col)
                        cell = ws.cell(dst_row, col)
                        if cell.__class__.__name__ == "MergedCell":
                            continue
                        if src_val is None:
                            if k > 0:
                                cell.value = None
                            continue
                        if cell.__class__.__name__ == "MergedCell":
                            continue
                        if isinstance(src_val, str):
                            new_val = src_val.replace(f"'{TEMPLATE_SHEET}'", f"'{sheet_name}'")
                            if k > 0:
                                # Thay row numbers: e.g. H50 → H56 khi k=1 (offset 6)
                                import re
                                for rel2 in range(BLOCK_SIZE):
                                    old_r = base_row + rel2
                                    new_r = module_base + rel2
                                    new_val = re.sub(
                                        r'(\$?[A-Z]+\$?)' + str(old_r) + r'(?!\d)',
                                        lambda m, nr=new_r: m.group(1) + str(nr),
                                        new_val
                                    )
                            cell.value = new_val
                        else:
                            cell.value = src_val
            else:
                # Clear block thừa — xóa ô có sheet refs, giữ col A (số TT) và col C (env name)
                for rel in range(BLOCK_SIZE):
                    dst_row = module_base + rel
                    for col in DATA_COLS:
                        cell = ws.cell(dst_row, col)
                        if cell.__class__.__name__ != "MergedCell":
                            cell.value = None


def build_excel_for_module(module_dir: Path, output_path: Path, tmpl: Path):
    """
    Tạo 1 file Excel từ tất cả JSON trong module_dir.
    Mỗi JSON file = 1 sheet.
    """
    json_files = sorted([f for f in module_dir.rglob("*.json") if f.name != "meta.json"])
    if not json_files:
        print(f"  Skipping {module_dir.name}: no JSON files")
        return

    shutil.copy2(str(tmpl), str(output_path))
    wb = openpyxl.load_workbook(str(output_path))

    for json_path in json_files:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)

        sheet_name = (data.get("module_name") or json_path.stem)[:31]

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            clear_data_rows(ws)
        elif TEMPLATE_SHEET in wb.sheetnames:
            ws = wb.copy_worksheet(wb[TEMPLATE_SHEET])
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        write_sheet(ws, data)
        print(f"    Sheet '{sheet_name}': {len(data.get('test_cases', []))} TCs")

    # Remove blank template sheet and wire Report formulas to real sheets
    non_system = [s for s in wb.sheetnames if s not in ("Cover", "Report")]
    real_sheets = [s for s in non_system if s != TEMPLATE_SHEET]
    if TEMPLATE_SHEET in wb.sheetnames and real_sheets:
        update_report_sheet(wb, real_sheets)
        del wb[TEMPLATE_SHEET]

    wb.save(str(output_path))
    print(f"  → Saved: {output_path}  ({len(json_files)} sheets)")


def json_to_excel(root_dir: str, output_dir: str, template_path: str = None):
    """
    Duyệt root_dir, mỗi sub-folder là 1 module lớn → tạo 1 file Excel riêng.
    Nếu root_dir chứa JSON trực tiếp (không có sub-folder), tạo 1 file Excel duy nhất
    lấy tên từ root_dir.

    Usage:
      python json_to_excel.py testcases/fdb-9999 output/
      → output/user.xlsx, output/student-management.xlsx ...
    """
    base = Path(root_dir)
    out  = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    if not base.exists():
        print(f"ERROR: Directory not found: {root_dir}")
        sys.exit(1)

    tmpl = Path(template_path) if template_path else DEFAULT_TPL
    if not tmpl.exists():
        print(f"ERROR: Template not found: {tmpl}")
        sys.exit(1)

    # Detect structure: sub-folders with JSON = multi-module; flat JSON = single module
    sub_dirs  = [d for d in sorted(base.iterdir()) if d.is_dir()]
    flat_json = [f for f in sorted(base.glob("*.json")) if f.name != "meta.json"]

    modules = []  # list of (module_dir, excel_filename)

    if sub_dirs:
        for d in sub_dirs:
            has_json = any(d.rglob("*.json"))
            if has_json:
                modules.append((d, f"{d.name}.xlsx"))

    # Also handle any flat JSON at root (treated as a module named after root_dir)
    if flat_json:
        modules.append((base, f"{base.name}.xlsx"))

    if not modules:
        print(f"No JSON files found in {root_dir}")
        sys.exit(1)

    for module_dir, excel_name in modules:
        print(f"\nModule: {module_dir.name}")
        build_excel_for_module(module_dir, out / excel_name, tmpl)

    print(f"\nDone. Output: {output_dir}")


if __name__ == "__main__":
    args = sys.argv[1:]
    tmpl_override = None
    if "--template" in args:
        i = args.index("--template")
        tmpl_override = args[i + 1]
        args = [a for j, a in enumerate(args) if j != i and j != i + 1]

    if len(args) < 2:
        print("Usage: python json_to_excel.py <root_dir> <output_dir> [--template <file.xlsx>]")
        sys.exit(1)

    json_to_excel(args[0], args[1], tmpl_override)
