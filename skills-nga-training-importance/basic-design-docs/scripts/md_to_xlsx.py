#!/usr/bin/env python3
"""Convert basic-design-docs reference Markdown templates to Excel (.xlsx)."""

from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...", file=sys.stderr)
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

MAX_COLS = 20
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
CODE_FILL = PatternFill("solid", fgColor="F2F2F2")
H1_FILL = PatternFill("solid", fgColor="4472C4")
H2_FILL = PatternFill("solid", fgColor="8EA9DB")
H3_FILL = PatternFill("solid", fgColor="B4C6E7")


def parse_table_row(line: str) -> list[str] | None:
    stripped = line.strip()
    if not stripped.startswith("|"):
        return None
    if stripped.replace("|", "").replace("-", "").replace(":", "").strip() == "":
        return None
    parts = [c.strip() for c in stripped.strip("|").split("|")]
    return parts


def is_separator_row(cells: list[str]) -> bool:
    return all(re.fullmatch(r":?-{1,}:?", c or "-") for c in cells)


def md_to_workbook(md_path: Path) -> Workbook:
    text = md_path.read_text(encoding="utf-8")
    lines = text.splitlines()

    wb = Workbook()
    ws = wb.active
    sheet_title = md_path.stem[:31]
    ws.title = sheet_title

    row = 1
    in_code = False
    code_lang = ""
    code_lines: list[str] = []
    i = 0

    def write_row(values: list[str], *, bold=False, fill=None, font_size=11, merge=False):
        nonlocal row
        if not values:
            return
        while len(values) < 1:
            values = [""]
        col_count = min(len(values), MAX_COLS)
        for c_idx in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c_idx, value=values[c_idx - 1])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
            if bold:
                cell.font = Font(bold=True, size=font_size, color="FFFFFF" if fill == H1_FILL else "000000")
            else:
                cell.font = Font(size=font_size)
            if fill:
                cell.fill = fill
        if merge and col_count > 1:
            ws.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=col_count,
            )
        row += 1

    def flush_code():
        nonlocal row, code_lines, code_lang
        if not code_lines:
            return
        label = f"[{code_lang or 'code'}]" if code_lang else "[code]"
        write_row([label], bold=True, fill=CODE_FILL, font_size=10)
        block = "\n".join(code_lines)
        write_row([block], fill=CODE_FILL, font_size=10)
        code_lines = []
        code_lang = ""
        row += 1

    while i < len(lines):
        line = lines[i]

        if line.strip().startswith("```"):
            if in_code:
                flush_code()
                in_code = False
            else:
                in_code = True
                code_lang = line.strip().removeprefix("```").strip()
                code_lines = []
            i += 1
            continue

        if in_code:
            code_lines.append(line)
            i += 1
            continue

        if line.strip() in ("---", "***", "___"):
            row += 1
            i += 1
            continue

        if line.startswith("# "):
            write_row([line[2:].strip()], bold=True, fill=H1_FILL, font_size=14, merge=True)
            i += 1
            continue
        if line.startswith("## "):
            write_row([line[3:].strip()], bold=True, fill=H2_FILL, font_size=12, merge=True)
            i += 1
            continue
        if line.startswith("### "):
            write_row([line[4:].strip()], bold=True, fill=H3_FILL, font_size=11, merge=True)
            i += 1
            continue
        if line.startswith("#### "):
            write_row([line[5:].strip()], bold=True, font_size=11, merge=True)
            i += 1
            continue

        cells = parse_table_row(line)
        if cells is not None:
            table_rows: list[list[str]] = []
            while i < len(lines):
                r = parse_table_row(lines[i])
                if r is None:
                    break
                if not is_separator_row(r):
                    table_rows.append(r)
                i += 1
            if table_rows:
                max_cols = max(len(r) for r in table_rows)
                for t_idx, trow in enumerate(table_rows):
                    padded = trow + [""] * (max_cols - len(trow))
                    is_header = t_idx == 0
                    write_row(padded[:MAX_COLS], bold=is_header, fill=HEADER_FILL if is_header else None)
                row += 1
            continue

        stripped = line.strip()
        if stripped.startswith("- ") or stripped.startswith("* "):
            write_row(["• " + stripped[2:].strip()])
            i += 1
            continue
        if re.match(r"^\d+\.\s", stripped):
            write_row([stripped])
            i += 1
            continue

        if stripped.startswith("**") and stripped.endswith("**") and ":" in stripped:
            write_row([stripped.replace("**", "")], bold=True)
            i += 1
            continue

        if stripped:
            write_row([stripped], merge=True)
        i += 1

    if in_code:
        flush_code()

    for col in range(1, min(ws.max_column or 1, MAX_COLS) + 1):
        letter = get_column_letter(col)
        max_len = 12
        for r in range(1, (ws.max_row or 1) + 1):
            val = ws.cell(row=r, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = "A2"
    return wb


def convert_all(source_dir: Path, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_resolved = output_dir.resolve()
    md_files = sorted(
        p
        for p in source_dir.rglob("*.md")
        if output_resolved not in p.resolve().parents and p.resolve().parent != output_resolved
    )
    created: list[Path] = []

    for md_path in md_files:
        wb = md_to_workbook(md_path)
        rel = md_path.relative_to(source_dir)
        out_name = rel.with_suffix(".xlsx").name if rel.parent == Path(".") else "-".join(rel.with_suffix(".xlsx").parts)
        out_path = output_dir / out_name
        wb.save(out_path)
        created.append(out_path)
        print(f"  {md_path.name} -> {out_path.name}")

    return created


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    references_dir = script_dir.parent / "references"
    output_dir = references_dir / "output template excel"

    if len(sys.argv) > 1:
        references_dir = Path(sys.argv[1]).resolve()
    if len(sys.argv) > 2:
        output_dir = Path(sys.argv[2]).resolve()

    print(f"Source: {references_dir}")
    print(f"Output: {output_dir}\n")
    created = convert_all(references_dir, output_dir)
    print(f"\nDone: {len(created)} file(s).")


if __name__ == "__main__":
    main()
